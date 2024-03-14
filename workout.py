import re, os
from datetime import datetime, date
import pandas as pd
import pygsheets
from openpyxl.utils import get_column_letter
from workoutobject import workout_list

from flask import request
from sqlalchemy import create_engine
from app import db, sql_keys

# Help received from ChatGPT refining this function
def exercise():
    # TODO: Add Server-side checks on entries
    # exercise_name = request.form.get("exercise")
    metric = request.form.get("load-metric")
    weights =[]
    sets = []
    numbers = ["one","two","three","four","five"]
    day = int(request.form.get("workout-day"))
    date = request.form.get("date")
    order = request.form.get("exercise")
    notes = request.form.get("notes")
    exercise_name = 'dummy'
    for workout in workout_list:
        if workout["Day"] == day:
            for exercise in workout["Exercises"]:
                if exercise["Order"] == order:
                    exercise_name = exercise["Name"]
                    pass
                else:
                    pass
        else:
            pass
    if order not in ['A', 'B', 'C', 'D', 'E', 'F']:
        exercise_name = order
        order = 'other'

    # Server-side checks to ensure Set One values aren't null
    try:
        set_one_reps_trial = float(request.form.get(f"set-one-reps"))
        set_one_weight_trial = float(request.form.get(f"set-one-weight"))
    except ValueError:
        return "ERROR"

    # Loop through all possible sets
    for i in numbers:
        set_value = request.form.get(f"set-{i}-reps")
        set_weight = request.form.get(f"set-{i}-weight")
        if set_value:
            sets.append(int(set_value))
            weights.append(set_weight)

    # Work out actual order of exercise entered
    current_order = 0
    actual_exercise_order = db.execute("SELECT ne.Day as Day, ne.Actual_Order as Actual_Order, ne.Date as Date FROM new_exercise ne JOIN (SELECT Day, MAX(Date) AS Max_Date FROM new_exercise GROUP BY Day) AS max_dates ON ne.Day = max_dates.Day AND ne.Date = max_dates.Max_Date;")
    for exercises in actual_exercise_order:
        if exercises['Day'] == day:
            if exercises['Actual_Order'] == None:
                exercises['Actual_Order'] = 0
            if current_order <= exercises['Actual_Order']:
                current_order = exercises['Actual_Order'] + 1


    # Upload into SQL tables
    db.execute("INSERT INTO new_exercise (Exercise, Actual_Order, Sheet_Order, Day, Date, Metric, Notes) VALUES (?, ?, ?, ?, ?, ?, ?)", exercise_name, current_order, order, day, date, metric, notes)

    # TODO: Investigate below two areas as potential source of NULL values - as it may be above row is working but either last_inserted_Id is failing, or sets[] is failing

    # Get the last inserted row ID (primary key)
    last_inserted_id = db.execute("SELECT LAST_INSERT_ID()")[0]['LAST_INSERT_ID()']

    # Upload into SQL tables
    for i in range(len(sets)):
        try:
            if sets[i]: # Had to insert numbers[i] into below due to MySQL - but these aren't inputted by user so should be safer
                db.execute(f"UPDATE new_exercise SET Set{numbers[i].capitalize()}_Reps = ?, Set{numbers[i].capitalize()}_Weight = ? WHERE ID = ?", sets[i], weights[i], last_inserted_id)
        except IndexError:
            pass

    return "success"

# Help received from ChatGPT refining this function
def weight_import():
    # Look for CSV file
    if 'weight_file' not in request.files:
        return 'No file part'
    file = request.files['weight_file']
    if file.filename == '':
        return 'No selected file'
    if file and file.filename.endswith('.csv'):
        # Process the CSV file
        pass
    else:
        return 'Please upload a CSV file'


    # Read CSV file
    if file:
        df = pd.read_csv(file)

    # Transpose CSV file into SQL
    # TODO: Be able to upload multiple weight files without deleting prior data
    
    
    engine = create_engine(f"mysql://{sql_keys['username']}:{sql_keys['password']}@BiggestChris.mysql.eu.pythonanywhere-services.com/BiggestChris$fitness")
    df.to_sql('weight', engine, if_exists='replace', index=False)
    engine.dispose()
    

    return "success"

# Taken using weight_import as a template
def food_import():
    # Look for CSV file
    if 'food_file' not in request.files:
        return 'No file part'
    file = request.files['food_file']
    if file.filename == '':
        return 'No selected file'
    if file and file.filename.endswith('.csv'):
        # Process the CSV file
        pass
    else:
        return 'Please upload a CSV file'


    # Read CSV file
    if file:
        df = pd.read_csv(file)

    # Transpose CSV file into SQL
    # TODO: Be able to upload multiple weight files without deleting prior data
    
    
    engine = create_engine(f"mysql://{sql_keys['username']}:{sql_keys['password']}@BiggestChris.mysql.eu.pythonanywhere-services.com/BiggestChris$fitness")
    df.to_sql('food', engine, if_exists='replace', index=False)
    engine.dispose()
    

    return "success"

# Help received from ChatGPT refining this function
def weight_export():

    # G-sheets authorisation
    # TODO: Need to load in JSON securely - having it in Repo and then uploading publicly created a security risk
        # json is now secured in PythonAnywhere - should be ok to switch repo back to public
    gc = pygsheets.authorize(service_file=r'/home/BiggestChris/Keys/g-sheets-for-python-a3ee6cd4d658.json')

    # Open the google spreadsheet (this has the key from the Greg Burns Fitness Sheet)
    sh = gc.open_by_key('1F_6EtWT68BO2EfY_dErs-fNKWiEMCj497Hs0MnI-HsY')

    #select the Daily Tracker worksheet
    wks = sh.worksheet_by_title('Daily Tracker')

    # TODO: Daily Tracker
        # Read Column C to pick up dates
    date_column = wks.get_col(3)
    # Above creates a list of values in the cells in that column

    # Need to compare to dates in Weight table
    dates = db.execute('SELECT `Time of Measurement`, `Weight(kg)` FROM weight')
    # Above creates a list of dictionaries, with 'Time of Measurement' and 'Weight(kg)' as the keys in those dictionaries

    fresh_dates = []

    for x in dates:
        date_object = datetime.strptime((x['Time of Measurement']), "%m/%d/%Y, %H:%M:%S")
        formatted_date = date_object.strftime("%d/%m/%y")
        fresh_dates.append({'date': formatted_date, 'weight': x['Weight(kg)']})

    # TODO: Change it so it loops through all rows in SQL table, rather than dates in G-sheet
    # Where they are matching - need to load in the weight into Column D
    for index, date in enumerate(date_column):
        # Ensure we skip the header row
        if index == 0:
            continue

        # Define the regex patterns
        pattern1 = r'^.{4}(\d{1}/\d{1}/\d{2})$'
        pattern2 = r'^.{4}(\d{1}/\d{2}/\d{2})$'
        pattern3 = r'^.{4}(\d{2}/\d{1}/\d{2})$'
        pattern4 = r'^.{4}(\d{2}/\d{2}/\d{2})$'

        match = re.match(pattern1, date) or re.match(pattern2, date) or re.match(pattern3, date) or re.match(pattern4, date)
        if match:
            # If the string matches the pattern
            # TODO: Fix error in how this is working - currently carrying over whole string
            extracted_part = match.group(1)
            # Extract the part representing the month
            day_part = extracted_part.split('/')[0]
            month_part = extracted_part.split('/')[1]
            # Check if the day or month part has only one digit
            if len(day_part) == 1:
                # Insert '0' before the single digit day
                extracted_part = '0' + extracted_part
            if len(month_part) == 1:
                # Insert '0' before the single digit month
                extracted_part = extracted_part[:3] + '0' + extracted_part[3:]

            # Iterate over fresh_dates to find a matching date
            for fresh_date in fresh_dates:
                if extracted_part == fresh_date['date']:
                    # Update the weight in Column D
                    wks.update_value(f"D{index+1}", fresh_date['weight'])
                    break  # Break the loop once we find a match

        else:
            pass

# Based on weight_export function
def food_export():

    # G-sheets authorisation
    # TODO: Need to load in JSON securely - having it in Repo and then uploading publicly created a security risk
        # json is now moved separately, but file location still in code which is a risk, repo moved to private for now until resolved
    gc = pygsheets.authorize(service_file=r'/home/BiggestChris/Keys/g-sheets-for-python-a3ee6cd4d658.json')

    # Open the google spreadsheet (this has the key from the Greg Burns Fitness Sheet)
    sh = gc.open_by_key('1F_6EtWT68BO2EfY_dErs-fNKWiEMCj497Hs0MnI-HsY')

    #select the Daily Tracker worksheet
    wks = sh.worksheet_by_title('Daily Tracker')

    # Read Column C to pick up dates
    date_column = wks.get_col(3)
    # Above creates a list of values in the cells in that column

    # Need to compare to dates in Food table
    dates = db.execute('SELECT Date, SUM(Calories) AS Calories, SUM(`Protein (g)`) AS Protein, SUM(`Carbohydrates (g)`) AS Carbohydrates, SUM(`Fat (g)`) AS Fat  FROM food GROUP BY Date')
    # Above creates a list of dictionaries, with 'Date', 'Calories', 'Protein', 'Carbohydrates', 'Fat' as the keys in those dictionaries, which sum up the total of those consumed per day

    fresh_dates = []

    for x in dates:
        date_object = datetime.strptime((x['Date']), "%Y-%m-%d")
        formatted_date = date_object.strftime("%d/%m/%y")
        fresh_dates.append({'date': formatted_date, 'calories': x['Calories'], 'protein': x['Protein'], 'carbohydrates': x['Carbohydrates'], 'fat': x['Fat']})

    # TODO: Change it so it loops through all rows in SQL table, rather than dates in G-sheet
    # Where they are matching - need to load in the weight into Column D
    for index, date in enumerate(date_column):
        # Ensure we skip the header row
        if index == 0:
            continue

        # Define the regex patterns
        pattern1 = r'^.{4}(\d{1}/\d{1}/\d{2})$'
        pattern2 = r'^.{4}(\d{1}/\d{2}/\d{2})$'
        pattern3 = r'^.{4}(\d{2}/\d{1}/\d{2})$'
        pattern4 = r'^.{4}(\d{2}/\d{2}/\d{2})$'

        match = re.match(pattern1, date) or re.match(pattern2, date) or re.match(pattern3, date) or re.match(pattern4, date)
        if match:
            # If the string matches the pattern
            # TODO: Fix error in how this is working - currently carrying over whole string
            extracted_part = match.group(1)
            # Extract the part representing the month
            day_part = extracted_part.split('/')[0]
            month_part = extracted_part.split('/')[1]
            # Check if the day or month part has only one digit
            if len(day_part) == 1:
                # Insert '0' before the single digit day
                extracted_part = '0' + extracted_part
            if len(month_part) == 1:
                # Insert '0' before the single digit month
                extracted_part = extracted_part[:3] + '0' + extracted_part[3:]

            # Iterate over fresh_dates to find a matching date
            for fresh_date in fresh_dates:
                if extracted_part == fresh_date['date']:
                    # Update the weight in Column D
                    wks.update_value(f"F{index+1}", fresh_date['calories'])
                    wks.update_value(f"G{index+1}", fresh_date['protein'])
                    wks.update_value(f"H{index+1}", fresh_date['carbohydrates'])
                    wks.update_value(f"I{index+1}", fresh_date['fat'])
                    break  # Break the loop once we find a match

        else:
            pass

# TODO: Refactor so much fewer API Calls made to GoogleSheets - as they get throttled
# Based on weight_export function
def exercise_export():

    # G-sheets authorisation
    # TODO: Need to load in JSON securely - having it in Repo and then uploading publicly created a security risk
        # json is now moved separately, but file location still in code which is a risk, repo moved to private for now until resolved
    gc = pygsheets.authorize(service_file=r'/home/BiggestChris/Keys/g-sheets-for-python-a3ee6cd4d658.json')

    # Open the google spreadsheet (this has the key from the Greg Burns Fitness Sheet)
    sh = gc.open_by_key('1F_6EtWT68BO2EfY_dErs-fNKWiEMCj497Hs0MnI-HsY')

    #select the Daily Tracker worksheet
    wks = sh.worksheet_by_title('Logbook')

    # Go through SQL table - then find where each exercise fits and fill in details
        # Find what date it is, then relate that to a Week
    exercises = db.execute('SELECT Date, Day, Sheet_Order, SetOne_Weight, SetOne_Reps, SetTwo_Weight, SetTwo_Reps, SetThree_Weight, SetThree_Reps, SetFour_Weight, SetFour_Reps FROM new_exercise')

    # Read the row and column from the worksheet - calling in here once to limit API Calls
    row_values = wks.get_row(13)
    column_values = wks.get_col(2)


    #ChatGPT helped refine this
    for exercise in exercises:

        # Find what date it is, then relate that to a Week
        #ChatGPT helped refine this
        exercise_date = datetime.strptime(exercise['Date'], '%Y-%m-%d').date()
        compare_date = date(2024, 1, 15)

        # Calculate the difference in days
        difference = (exercise_date - compare_date).days
        exercise_week = difference // 7 + 1


        # Locate relevant columns for that Week on the G-sheet
        # Find the index of the first cell in the row for that Week
        week_column_index = next((index for index, value in enumerate(row_values) if value == f'WEEK {exercise_week}'), None)

        # If column_index is None, it means an empty cell was found
        if week_column_index is None:
            pass

        else:
        # Find what workout Day it is in the table, locate relevant set of rows on the G-sheet
            day_row_index = next((index for index, value in enumerate(column_values) if value == f'DAY {exercise["Day"]}'), None)

        # Then look up workout Order, find where that is in Column B in the relevant set of rows informed by Day
            if day_row_index is None:
                pass
            else:
                values_to_search = column_values[(day_row_index + 1):(day_row_index + 9)]


                exercise_value_index = next((index for index, value in enumerate(values_to_search) if value == exercise['Sheet_Order']), None)

                if exercise_value_index is None:
                    pass
                else:
                    # New index gives where to look
                    adjusted_index = day_row_index + exercise_value_index + 1

                    # Look up set 1 load and set 1 reps - put into first two columns in relevant Week columns - remember, get_column_letter isn't zero-indexed
                    # wks.update_value(f"{get_column_letter(week_column_index + 1)}{adjusted_index + 1}", exercise["SetOne_Weight"])
                    # wks.update_value(f"{get_column_letter(week_column_index + 2)}{adjusted_index + 1}", exercise["SetOne_Reps"])

                    # Do the same for sets 2,3,4 and second two, third two, fourth two columns of relevant Week columns
                    for x in [(1,'One'),(2,'Two'),(3,'Three'),(4,'Four')]:
                        wks.update_value(f"{get_column_letter(week_column_index + (2 * (x[0] - 1)) + 1)}{adjusted_index + 1}", exercise[f"Set{x[1]}_Weight"])
                        wks.update_value(f"{get_column_letter(week_column_index + (2 * (x[0] - 1)) + 2)}{adjusted_index + 1}", exercise[f"Set{x[1]}_Reps"])
                    # Repeat for all rows in the SQL table


    else:
        pass

# Need to retrieve latest workouts from SQL and store in an object
def retrieve_workout():

    # ChatGPT helped with this query - Gets workout exercises from last day BEFORE TODAY, assumes user will only use input on day before current day
    last_workout = db.execute('SELECT ne.Day, ne.Actual_Order, ne.Sheet_Order, ne.Exercise, ne.SetOne_Weight, ne.SetOne_Reps, ne.SetTwo_Weight, ne.SetTwo_Reps, ne.SetThree_Weight, ne.SetThree_Reps, ne.SetFour_Weight, ne.SetFour_Reps, ne.SetFive_Weight, ne.SetFive_Reps FROM new_exercise ne JOIN (SELECT Day, Sheet_Order, MAX(Date) AS Max_Date FROM new_exercise WHERE date < DATE(NOW()) GROUP BY Day) AS max_dates ON ne.Day = max_dates.Day AND ne.Date = max_dates.Max_Date;')
    # last_workout = db.execute('SELECT ne.Day, ne.Actual_Order, ne.Sheet_Order, ne.Exercise, ne.SetOne_Weight, ne.SetOne_Reps, ne.SetTwo_Weight, ne.SetTwo_Reps, ne.SetThree_Weight, ne.SetThree_Reps, ne.SetFour_Weight, ne.SetFour_Reps, ne.SetFive_Weight, ne.SetFive_Reps FROM new_exercise ne JOIN (SELECT Day, Sheet_Order, MAX(Date) AS Max_Date FROM new_exercise GROUP BY Day) AS max_dates ON ne.Day = max_dates.Day AND ne.Date = max_dates.Max_Date;')
    # last_workout = db.execute('SELECT ne.Day, ne.Actual_Order, ne.Sheet_Order, ne.Exercise, ne.SetOne_Weight, ne.SetOne_Reps, ne.SetTwo_Weight, ne.SetTwo_Reps, ne.SetThree_Weight, ne.SetThree_Reps, ne.SetFour_Weight, ne.SetFour_Reps, ne.SetFive_Weight, ne.SetFive_Reps FROM new_exercise ne JOIN (SELECT Day, Sheet_Order, MAX(Date) AS Max_Date FROM new_exercise GROUP BY Day, Sheet_Order) AS max_dates ON ne.Day = max_dates.Day AND ne.Date = max_dates.Max_Date AND ne.Sheet_Order = max_dates.Sheet_Order;')
    return last_workout
